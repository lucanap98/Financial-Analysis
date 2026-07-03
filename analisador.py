import math
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Lista única de campos usada tanto pelo template quanto pela digitação manual.
# Tupla: (label, tipo) — tipo None indica cabeçalho de seção.
CAMPOS = [
    ("IDENTIFICAÇÃO", None),
    ("Nome da Empresa", "texto"),
    ("Período (ex: 2024)", "texto"),
    ("DRE", None),
    ("Receita Líquida", "numero"),
    ("Custo dos Produtos Vendidos", "numero"),
    ("Lucro Bruto", "numero"),
    ("Despesas Operacionais", "numero"),
    ("EBIT", "numero"),
    ("Depreciação e Amortização", "numero"),
    ("Despesas Financeiras", "numero"),
    ("Lucro Líquido", "numero"),
    ("BALANÇO — ATIVO", None),
    ("Caixa e Equivalentes", "numero"),
    ("Contas a Receber", "numero"),
    ("Estoques", "numero"),
    ("Ativo Circulante", "numero"),
    ("Ativo Total", "numero"),
    ("BALANÇO — PASSIVO E PL", None),
    ("Fornecedores", "numero"),
    ("Passivo Circulante", "numero"),
    ("Dívida Curto Prazo", "numero"),
    ("Dívida Longo Prazo", "numero"),
    ("Patrimônio Líquido", "numero"),
    ("FLUXO DE CAIXA", None),
    ("Fluxo de Caixa Operacional", "numero"),
    ("CapEx", "numero"),
]


def safe_div(numerador, denominador):
    """Divisão protegida: retorna NaN quando o denominador é zero ou ausente.

    Demonstrações reais têm edge cases (PL negativo é reportado normalmente,
    mas receita zero ou despesa financeira zero quebrariam a divisão).
    """
    try:
        if denominador == 0 or pd.isna(denominador) or pd.isna(numerador):
            return float("nan")
        return numerador / denominador
    except TypeError:
        return float("nan")


def fmt(valor, formato, sufixo=""):
    """Formata o índice ou exibe 'n/d' quando o cálculo não foi possível."""
    if isinstance(valor, float) and math.isnan(valor):
        return "n/d"
    return format(valor, formato) + sufixo


# Essa função gera um template Excel com os campos necessários para a análise
# financeira. Os cabeçalhos de seção são destacados em azul e os campos de
# dados em amarelo claro. O usuário preenche a coluna B e usa a opção 2
# para carregar os dados.
def gerar_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    for row, (label, tipo) in enumerate(CAMPOS, start=1):
        celula_a = ws.cell(row=row, column=1, value=label)
        celula_b = ws.cell(row=row, column=2)

        if tipo is None:
            celula_a.font = Font(bold=True, color="FFFFFF")
            celula_a.fill = PatternFill("solid", start_color="2F5496", fgColor="2F5496")
            celula_b.fill = PatternFill("solid", start_color="2F5496", fgColor="2F5496")
        else:
            celula_b.fill = PatternFill("solid", start_color="FFFACD", fgColor="FFFACD")

    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_financeiro.xlsx")
    wb.save(caminho)
    print()
    print(f"Template gerado com sucesso: {caminho}")
    print("Preencha a coluna B e use a opção 2 para carregar.")


# Essa função carrega os dados de um arquivo CSV. O arquivo deve ter duas
# colunas: "campo" e "valor" (separadas por ponto e vírgula).
def carregar_csv():
    caminho = input("Caminho do arquivo CSV: ").strip().strip('"')
    df = pd.read_csv(caminho, sep=";")

    dados = {}
    for _, linha in df.iterrows():
        dados[linha["campo"]] = linha["valor"]

    exibir_dados(dados)
    return dados


# Essa função carrega os dados do template Excel preenchido. Linhas de
# cabeçalho de seção (coluna B vazia) são ignoradas automaticamente.
def carregar_excel():
    caminho = input("Caminho do arquivo Excel: ").strip().strip('"')
    df = pd.read_excel(caminho, header=None)

    dados = {}
    for _, linha in df.iterrows():
        label = linha[0]
        valor = linha[1]

        if pd.isna(valor) or pd.isna(label):
            continue

        dados[label] = valor

    exibir_dados(dados)
    return dados


# Digitação manual: percorre a mesma lista de campos do template,
# pedindo cada valor. Enter em branco pula o campo (vira NaN).
def digitar_manual():
    print()
    print("Digite os valores. Enter em branco pula o campo.")
    dados = {}

    for label, tipo in CAMPOS:
        if tipo is None:
            print(f"\n--- {label} ---")
            continue

        entrada = input(f"  {label}: ").strip()

        if entrada == "":
            dados[label] = float("nan")
        elif tipo == "numero":
            try:
                dados[label] = float(entrada.replace(".", "").replace(",", "."))
            except ValueError:
                print("    Valor inválido, campo ignorado.")
                dados[label] = float("nan")
        else:
            dados[label] = entrada

    exibir_dados(dados)
    return dados


def exibir_dados(dados):
    print()
    print("Dados carregados:")
    for campo, valor in dados.items():
        print(f"  {campo}: {valor}")


def calcular_indices(dados):
    g = dados.get  # atalho: .get() evita KeyError se algum campo faltar

    # Rentabilidade
    margem_bruta   = safe_div(g("Lucro Bruto"), g("Receita Líquida"))
    margem_liquida = safe_div(g("Lucro Líquido"), g("Receita Líquida"))
    margem_ebit    = safe_div(g("EBIT"), g("Receita Líquida"))
    ebitda         = (g("EBIT") or 0) + (g("Depreciação e Amortização") or 0)
    margem_ebitda  = safe_div(ebitda, g("Receita Líquida"))
    roa            = safe_div(g("Lucro Líquido"), g("Ativo Total"))
    roe            = safe_div(g("Lucro Líquido"), g("Patrimônio Líquido"))

    # Liquidez
    liquidez_corrente = safe_div(g("Ativo Circulante"), g("Passivo Circulante"))
    liquidez_seca     = safe_div((g("Ativo Circulante") or 0) - (g("Estoques") or 0), g("Passivo Circulante"))
    liquidez_imediata = safe_div(g("Caixa e Equivalentes"), g("Passivo Circulante"))

    # Endividamento
    divida_total    = (g("Dívida Curto Prazo") or 0) + (g("Dívida Longo Prazo") or 0)
    divida_liquida  = divida_total - (g("Caixa e Equivalentes") or 0)
    de_ratio        = safe_div(divida_total, g("Patrimônio Líquido"))
    cobertura_juros = safe_div(g("EBIT"), g("Despesas Financeiras"))
    divida_ebitda   = safe_div(divida_liquida, ebitda)

    # Eficiência (base 360 dias)
    pmr = safe_div((g("Contas a Receber") or 0) * 360, g("Receita Líquida"))
    pme = safe_div((g("Estoques") or 0) * 360, g("Custo dos Produtos Vendidos"))
    pmp = safe_div((g("Fornecedores") or 0) * 360, g("Custo dos Produtos Vendidos"))
    ciclo_financeiro = pmr + pme - pmp  # NaN se qualquer prazo for NaN

    # Fluxo de caixa
    fcl             = (g("Fluxo de Caixa Operacional") or 0) - (g("CapEx") or 0)
    qualidade_lucro = safe_div(g("Fluxo de Caixa Operacional"), g("Lucro Líquido"))

    print()
    print("=== RENTABILIDADE ===")
    print(f"  Margem Bruta:    {fmt(margem_bruta, '.1%')}")
    print(f"  Margem EBIT:     {fmt(margem_ebit, '.1%')}")
    print(f"  Margem EBITDA:   {fmt(margem_ebitda, '.1%')}")
    print(f"  Margem Líquida:  {fmt(margem_liquida, '.1%')}")
    print(f"  ROA:             {fmt(roa, '.1%')}")
    print(f"  ROE:             {fmt(roe, '.1%')}")

    print()
    print("=== LIQUIDEZ ===")
    print(f"  Liquidez Corrente:  {fmt(liquidez_corrente, '.2f', 'x')}")
    print(f"  Liquidez Seca:      {fmt(liquidez_seca, '.2f', 'x')}")
    print(f"  Liquidez Imediata:  {fmt(liquidez_imediata, '.2f', 'x')}")

    print()
    print("=== ENDIVIDAMENTO ===")
    print(f"  Dívida / PL:           {fmt(de_ratio, '.2f', 'x')}")
    print(f"  Cobertura de Juros:    {fmt(cobertura_juros, '.2f', 'x')}")
    print(f"  Dívida Líq. / EBITDA:  {fmt(divida_ebitda, '.2f', 'x')}")

    print()
    print("=== EFICIÊNCIA ===")
    print(f"  PMR (recebimento):  {fmt(pmr, '.1f', ' dias')}")
    print(f"  PME (estoque):      {fmt(pme, '.1f', ' dias')}")
    print(f"  PMP (pagamento):    {fmt(pmp, '.1f', ' dias')}")
    print(f"  Ciclo Financeiro:   {fmt(ciclo_financeiro, '.1f', ' dias')}")

    print()
    print("=== FLUXO DE CAIXA ===")
    print(f"  FCL (Caixa Livre):   R$ {fcl:,.0f}")
    print(f"  Qualidade do Lucro:  {fmt(qualidade_lucro, '.2f', 'x')}")


def menu():
    print("Analisador de Demonstrações Financeiras")
    print("=========================================")
    print()
    print("Como deseja inserir os dados?")
    print("  0 - Gerar template Excel")
    print("  1 - Arquivo CSV")
    print("  2 - Arquivo Excel")
    print("  3 - Digitação manual")
    print()

    opcao = input("Digite o número da opção: ").strip()

    dados = None

    if opcao == "0":
        gerar_template()
    elif opcao == "1":
        print("Vamos carregar um arquivo CSV!")
        dados = carregar_csv()
    elif opcao == "2":
        print("Vamos carregar um arquivo Excel!")
        dados = carregar_excel()
    elif opcao == "3":
        dados = digitar_manual()
    else:
        print("Opção inválida. Por favor escolha 0, 1, 2 ou 3.")

    if dados is not None:
        print()
        print("Dados carregados com sucesso! Total de campos:", len(dados))
        print("Calculando índices financeiros...")
        calcular_indices(dados)


if __name__ == "__main__":
    menu()
