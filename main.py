import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

#Essa função gera um template Excel com os campos necessários para a análise financeira. Os campos de identificação são destacados em azul, enquanto os campos de dados são destacados em amarelo claro. O usuário pode preencher a coluna B e depois usar a opção 2 para carregar os dados.
def gerar_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    campos = [
        ("IDENTIFICAÇÃO", None),
        ("Nome da Empresa", "texto"),
        ("Período (ex: 2024)", "texto"),
        ("DRE", None),
        ("Receita Líquida", "numero"),
        ("Custo dos Produtos Vendidos", "numero"),
        ("Lucro Bruto", "numero"),
        ("Despesas Operacionais", "numero"),
        ("EBIT", "numero"),
        ("Depreciação e Amortização", "numero"),
        ("Despesas Financeiras", "numero"),
        ("Lucro Líquido", "numero"),
        ("BALANÇO — ATIVO", None),
        ("Caixa e Equivalentes", "numero"),
        ("Contas a Receber", "numero"),
        ("Estoques", "numero"),
        ("Ativo Circulante", "numero"),
        ("Ativo Total", "numero"),
        ("BALANÇO — PASSIVO E PL", None),
        ("Fornecedores", "numero"),
        ("Passivo Circulante", "numero"),
        ("Dívida Curto Prazo", "numero"),
        ("Dívida Longo Prazo", "numero"),
        ("Patrimônio Líquido", "numero"),
        ("FLUXO DE CAIXA", None),
        ("Fluxo de Caixa Operacional", "numero"),
        ("CapEx", "numero"),
    ]

    for row, (label, tipo) in enumerate(campos, start=1):
        celula_a = ws.cell(row=row, column=1, value=label)
        celula_b = ws.cell(row=row, column=2)

        if tipo is None:
            celula_a.font = Font(bold=True, color="FFFFFF")
            celula_a.fill = PatternFill("solid", start_color="2F5496", fgColor="2F5496")
            celula_b.fill = PatternFill("solid", start_color="2F5496", fgColor="2F5496")
        else:
            celula_b.fill = PatternFill("solid", start_color="FFFACD", fgColor="FFFACD")

    caminho = os.path.join(os.path.dirname(__file__), "template_financeiro.xlsx")
    wb.save(caminho)
    print()
    print(f"Template gerado com sucesso: {caminho}")
    print("Preencha a coluna B e use a opção 2 para carregar.")

#Essa função carrega os dados de um arquivo CSV. O arquivo deve ter duas colunas: "campo" e "valor". A função lê o CSV, armazena os dados em um dicionário e exibe os campos e valores carregados.
def carregar_csv():
    caminho = input("Anexe o arquivo CSV: ")
    df = pd.read_csv(caminho, sep=";")

    dados = {}
    for _, linha in df.iterrows():
        dados[linha["campo"]] = linha["valor"]

    print()
    print("Dados carregados:")
    for campo, valor in dados.items():
        print(f"  {campo}: {valor}")

    return dados


def carregar_excel():
    caminho = input("Anexe o arquivo Excel: ")
    df = pd.read_excel(caminho, header=None)

    dados = {}
    for _, linha in df.iterrows():
        label = linha[0]
        valor = linha[1]

        if pd.isna(valor) or isinstance(label, float):
            continue

        dados[label] = valor

    print()
    print("Dados carregados:")
    for campo, valor in dados.items():
        print(f"  {campo}: {valor}")

    return dados


def digitar_manual():
    print("Digite os dados manualmente.")


def calcular_indices(dados):
    # Rentabilidade
    margem_bruta   = dados["Lucro Bruto"] / dados["Receita Líquida"]
    margem_liquida = dados["Lucro Líquido"] / dados["Receita Líquida"]
    margem_ebit    = dados["EBIT"] / dados["Receita Líquida"]
    ebitda         = dados["EBIT"] + dados["Depreciação e Amortização"]
    margem_ebitda  = ebitda / dados["Receita Líquida"]
    roa            = dados["Lucro Líquido"] / dados["Ativo Total"]
    roe            = dados["Lucro Líquido"] / dados["Patrimônio Líquido"]

    # Liquidez
    liquidez_corrente = dados["Ativo Circulante"] / dados["Passivo Circulante"]
    liquidez_seca     = (dados["Ativo Circulante"] - dados["Estoques"]) / dados["Passivo Circulante"]
    liquidez_imediata = dados["Caixa e Equivalentes"] / dados["Passivo Circulante"]

    # Endividamento
    divida_total   = dados["Dívida Curto Prazo"] + dados["Dívida Longo Prazo"]
    divida_liquida = divida_total - dados["Caixa e Equivalentes"]
    de_ratio       = divida_total / dados["Patrimônio Líquido"]
    cobertura_juros = dados["EBIT"] / dados["Despesas Financeiras"]
    divida_ebitda  = divida_liquida / ebitda

    # Eficiência
    pmr              = (dados["Contas a Receber"] * 360) / dados["Receita Líquida"]
    pme              = (dados["Estoques"] * 360) / dados["Custo dos Produtos Vendidos"]
    pmp              = (dados["Fornecedores"] * 360) / dados["Custo dos Produtos Vendidos"]
    ciclo_financeiro = pmr + pme - pmp

    # Fluxo de caixa
    fcl            = dados["Fluxo de Caixa Operacional"] - dados["CapEx"]
    qualidade_lucro = dados["Fluxo de Caixa Operacional"] / dados["Lucro Líquido"]

    print()
    print("=== RENTABILIDADE ===")
    print(f"  Margem Bruta:    {margem_bruta:.1%}")
    print(f"  Margem EBIT:     {margem_ebit:.1%}")
    print(f"  Margem EBITDA:   {margem_ebitda:.1%}")
    print(f"  Margem Líquida:  {margem_liquida:.1%}")
    print(f"  ROA:             {roa:.1%}")
    print(f"  ROE:             {roe:.1%}")

    print()
    print("=== LIQUIDEZ ===")
    print(f"  Liquidez Corrente:  {liquidez_corrente:.2f}x")
    print(f"  Liquidez Seca:      {liquidez_seca:.2f}x")
    print(f"  Liquidez Imediata:  {liquidez_imediata:.2f}x")

    print()
    print("=== ENDIVIDAMENTO ===")
    print(f"  Dívida / PL:           {de_ratio:.2f}x")
    print(f"  Cobertura de Juros:    {cobertura_juros:.2f}x")
    print(f"  Dívida Líq. / EBITDA:  {divida_ebitda:.2f}x")

    print()
    print("=== EFICIÊNCIA ===")
    print(f"  PMR (recebimento):  {pmr:.1f} dias")
    print(f"  PME (estoque):      {pme:.1f} dias")
    print(f"  PMP (pagamento):    {pmp:.1f} dias")
    print(f"  Ciclo Financeiro:   {ciclo_financeiro:.1f} dias")

    print()
    print("=== FLUXO DE CAIXA ===")
    print(f"  FCL (Caixa Livre):   R$ {fcl:,.0f}")
    print(f"  Qualidade do Lucro:  {qualidade_lucro:.2f}x")


def menu():
    print("Analisador de Demonstrações Financeiras")
    print("=========================================")
    print()
    print("Como deseja inserir os dados?")
    print("  0 - Gerar template Excel")
    print("  1 - Arquivo CSV")
    print("  2 - Arquivo Excel")
    print("  3 - Digitação manual")
    print()

    opcao = input("Digite o número da opção: ")

    dados = None

    if opcao == "0":
        gerar_template()
    elif opcao == "1":
        print("Vamos carregar um arquivo CSV!")
        dados = carregar_csv()
    elif opcao == "2":
        print("Vamos carregar um arquivo Excel!")
        dados = carregar_excel()
    elif opcao == "3":
        print("Vamos digitar os dados manualmente!")
        digitar_manual()
    else:
        print("Opção inválida. Por favor escolha 0, 1, 2 ou 3.")

    if dados is not None:
        print()
        print("Dados carregados com sucesso! Total de campos:", len(dados))
        print("Calculando índices financeiros...")
        calcular_indices(dados)


menu()